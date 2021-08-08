"""
This bot performs a bit of web scraping to get current values
of several currency pairs and financial instruments.
"""

import time
from datetime import datetime, timedelta

import praw
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

import config

HEADERS = {
    "User-Agent": "FinanceBot v0.3"}

INVESTING_DICT = {
    "USD/MXN": "https://mx.investing.com/currencies/usd-mxn",
    "EUR/MXN": "https://mx.investing.com/currencies/eur-mxn",
    "GBP/MXN": "https://mx.investing.com/currencies/gbp-mxn",
    "BTC/MXN": "https://www.investing.com/crypto/bitcoin/btc-mxn",
    "IPC (BMV)": "https://mx.investing.com/indices/ipc"
}

BANXICO1_URL = "https://www.banxico.org.mx/SieInternet/consultarDirectorioInternetAction.do?sector=22&accion=consultarCuadro&idCuadro=CF107&locale=es&formatoXLS.x=1&fechaInicio={}&fechaFin={}"
BANXICO2_URL = "https://www.banxico.org.mx/SieInternet/consultarDirectorioInternetAction.do?accion=consultarCuadro&idCuadro=CF114&formatoXLS.x=1&fechaInicio={}&fechaFin={}"


def init_bot():
    """Inits the bot."""

    # We create the Reddit instance.
    reddit = praw.Reddit(client_id=config.APP_ID, client_secret=config.APP_SECRET,
                         user_agent=config.USER_AGENT, username=config.REDDIT_USERNAME,
                         password=config.REDDIT_PASSWORD)

    # Load the pre-existing sirebar text.
    sidebar_text = open("sidebar.txt", "r", encoding="utf-8").read()

    # Start the Markdown table with 3 columns.
    table_text = """\n\n| | | |\n| --- | --- | --- |\n"""

    # We iterate over INVESTING_DICT and call the same function.
    for k, v in INVESTING_DICT.items():

        temp_data = get_investing_data(k, v)

        # Add the data to the Markdown table.
        table_text += "| {} | {} | {} |\n".format(
            temp_data[0], temp_data[1], temp_data[2])

        time.sleep(1)

    # We add the rest of financial instruments.
    for item in get_cetes():
        table_text += "| {} | | {} |\n".format(item[0], item[1])

    # Prepare the footer with the current date and time.
    now = datetime.now()
    footer = "\nÚltima actualización: {:%d-%m-%Y a las %H:%M:%S}".format(now)

    # Update the sidebar on old Reddit.
    reddit.subreddit(
        config.SUBREDDIT).wiki["config/sidebar"].edit(sidebar_text + table_text + footer)

    # Update a sidebar text widget on new Reddit.
    for widget in reddit.subreddit(config.SUBREDDIT).widgets.sidebar:

        if widget.shortName == "Indicadores Financieros":
            widget.mod.update(text=table_text + footer)
            break


def get_investing_data(name, url):
    """Performs web scraping and gets 3 values from investing.com

    Parameters
    ----------
    name : str
        The name of the currency pair.

    url : str
        The url to scrape.

    Returns
    -------
    tuple
         A Tuple containing 3 values.

    """

    with requests.get(url, headers=HEADERS) as response:

        soup = BeautifulSoup(response.text, "html.parser")

        try:
            latest_data = soup.find(
                "div", {"class": "top bold inlineblock"}).text.strip().split()

            price = latest_data[0]
            percentage = latest_data[2]
        except:
            price = soup.find(
                "span", {"data-test": "instrument-price-last"}).text.strip()

            percentage = soup.find(
                "span", {"data-test": "instrument-price-change-percent"}).text.strip()[1:-1]

        return (name, price, percentage)


def get_cetes():
    """Gets data from Banxico Excel archives.
    It first downloads the Excel files and then reads
    cell values from the specified rows.

    Returns
    -------
    list
        A list of several financial instruments.

    """

    # The Excel files are behind a simple GET request.
    # Two of the parameters are UNIX timestamps.
    now = datetime.now()
    now_ts = int(now.timestamp()) * 1000

    last_21_days = now - timedelta(days=21)
    last_21_days_ts = int(last_21_days.timestamp()) * 1000

    last_120_days = now - timedelta(days=120)
    last_120_days_ts = int(last_120_days.timestamp()) * 1000

    # With our timestamps ready we download both files.
    with requests.get(BANXICO1_URL.format(last_21_days_ts, now_ts)) as response:
        with open("./banxico1.xlsx", "wb") as temp_file:
            temp_file.write(response.content)

    with requests.get(BANXICO2_URL.format(last_120_days_ts, now_ts)) as response:
        with open("./banxico2.xlsx", "wb") as temp_file:
            temp_file.write(response.content)

    # We open both files and start extracting the values we need.
    book1 = load_workbook("banxico1.xlsx")
    sheet1 = book1.worksheets[0]

    book2 = load_workbook("banxico2.xlsx")
    sheet2 = book2.worksheets[0]

    data_list = list()

    data_list.append(["CETES 1 mes", "+{}%".format(find_value(sheet1, 16))])
    data_list.append(["CETES 3 meses", "+{}%".format(find_value(sheet1, 20))])
    data_list.append(["CETES 6 meses", "+{}%".format(find_value(sheet1, 24))])
    data_list.append(["CETES 1 año", "+{}%".format(find_value(sheet1, 28))])

    data_list.append(["BONOS 3 años", "+{}%".format(find_value(sheet2, 29))])
    data_list.append(["BONOS 5 años", "+{}%".format(find_value(sheet2, 30))])
    data_list.append(["BONOS 10 años", "+{}%".format(find_value(sheet2, 32))])
    data_list.append(["BONOS 20 años", "+{}%".format(find_value(sheet2, 33))])
    data_list.append(["BONOS 30 años", "+{}%".format(find_value(sheet2, 34))])

    data_list.append(
        ["UDIBONOS 3 años", "+{}% (más inflación)".format(find_value(sheet2, 23))])

    data_list.append(
        ["UDIBONOS 10 años", "+{}% (más inflación)".format(find_value(sheet2, 25))])

    data_list.append(
        ["UDIBONOS 20 años", "+{}% (más inflación)".format(find_value(sheet2, 26))])

    data_list.append(
        ["UDIBONOS 30 años", "+{}% (más inflación)".format(find_value(sheet2, 27))])

    return data_list


def find_value(sheet, row_number):
    """Finds the best available value.

    Parameters
    ---------
    sheet : workbook.sheet
        An Excel sheet.

    row_number : int
        The row where the value is located.

    Returns
    -------
    string
        The cell value that wasn't 'N/E' (not eligible).

    """

    # We first try looking in the fifth column and keep falling
    # back until we go to the third column.
    if sheet.cell(row=row_number, column=5).value != "N/E":
        return sheet.cell(row=row_number, column=5).value
    elif sheet.cell(row=row_number, column=4).value != "N/E":
        return sheet.cell(row=row_number, column=4).value
    elif sheet.cell(row=row_number, column=3).value != "N/E":
        return sheet.cell(row=row_number, column=3).value
    elif sheet.cell(row=row_number, column=2).value != "N/E":
        return sheet.cell(row=row_number, column=2).value


if __name__ == "__main__":

    init_bot()
